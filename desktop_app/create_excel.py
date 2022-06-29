import openpyxl
from openpyxl import load_workbook
import textwrap

from qr_code_make import create_qr

static_path = "/home/f1r980y/PycharmProjects/N9XT2/desktop_app/static/"

# cross-platform @decorator need
# def order_print():
#     os.startfile(f"{static_path}excel_files/examples/order.xlsx", "print")
#
#
# def warranty_print():
#     os.startfile(f"{static_path}excel_files/examples/warranty.xlsx", "print")


def paste_to_order(model, name, phone_num, package, breakage, date, slug):
    url = f'http://172.20.10.8:8000/client_card/{slug}/'
    create_qr(url)
    wb = load_workbook(f"{static_path}excel_files/examples/order.xlsx")
    ws = wb.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/client.png')
    img.anchor = 'K11'
    ws.add_image(img)
    short_date = date.strftime("%d.%m.%Y")
    longest_date = date.strftime("%d.%m.%Y %H:%M")
    string_width = 25
    ws["C10"], ws["L10"], ws["C11"], ws["L11"], ws["C13"], ws["L13"], ws["C14"] =\
        model, model, package, package, name, name, phone_num
    ws["C17"], ws["M18"] = short_date, short_date
    breakage = textwrap.fill(breakage, string_width).split("\n")
    ws["C12"], ws["L12"] = breakage[0], breakage[0]
    #wb.save(f"{static_path}excel_files/examples/order.xlsx")
    wb.save(f"{static_path}excel_files/sawed_xlsx/{name}_{longest_date}.xlsx")


def paste_to_warranty(client_id, break_fix, date, price, warranty):
    wb = load_workbook(f"{static_path}excel_files/examples/warranty.xlsx")
    ws = wb.worksheets[0]
    short_date = date.strftime("%d.%m.%Y")
    string_width = 30
    ws["d10"] = client_id
    break_fix = textwrap.fill(break_fix, string_width).split("\n")
    ws["A11"] = break_fix[0]
    if len(break_fix) > 1:
        ws["A12"] = break_fix[1]
    else:
        ws["A12"] = ""
    if len(break_fix) > 2:
        ws["A13"] = break_fix[2]
    else:
        ws["A13"] = ""
    ws["B14"] = warranty
    ws["C15"] = price + ".00грн."
    ws["B16"] = short_date
    wb.save(f"{static_path}excel_files/examples/warranty.xlsx")

