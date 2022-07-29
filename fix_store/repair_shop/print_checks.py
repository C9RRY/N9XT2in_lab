import openpyxl
from openpyxl import load_workbook
import textwrap
import qrcode
from pathlib import Path


static_path = Path(__file__).resolve().parent.parent


def create_qr(url):
    qr = qrcode.QRCode(version=2, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=4, border=1)
    qr.add_data(url)
    qr.make(fit=True)
    image = qr.make_image()
    image.save(f"{static_path}/media/img/client.png")


def paste_to_order(data):
    model, package, breakage, name, phone_num, date, slug =\
        data[2], data[3], data[4], data[5], data[6], data[7], data[9]
    url = f'http://172.20.10.8:8000/client_card/{slug}/'
    create_qr(url)
    wb = load_workbook(f"{static_path}/media/excel_files/examples/order.xlsx")
    ws = wb.worksheets[0]
    short_date = date.strftime("%d.%m.%Y")
    longest_date = date.strftime("%d.%m.%Y %H:%M")
    string_width = 25
    ws["A2"], ws["A19"], ws["A11"], ws["A28"], ws["A13"] =\
        model, model, name, name, phone_num
    ws["A15"], ws["A33"] = short_date, short_date
    package = textwrap.fill(package, string_width).split("\n")
    ws["A4"], ws["A21"] = package[0], package[0]
    if len(package) >= 2:
        ws["A5"], ws["A22"] = package[1], package[1]
    breakage = textwrap.fill(breakage, string_width).split("\n")
    ws["A7"], ws["A24"] = breakage[0], breakage[0]
    if len(breakage) >= 2:
        ws["A8"], ws["A25"] = breakage[1], breakage[1]
    if len(breakage) >= 3:
        ws["A9"], ws["A26"] = breakage[2], breakage[2]
    wb.save(f"{static_path}/media/excel_files/sawed_xlsx/{name}_{longest_date}.xlsx")
    return f'media/excel_files/sawed_xlsx/{name}_{longest_date}.xlsx'


def paste_to_warranty(data):

    model, name, phone_num, date, slug, break_fix, price, warranty =\
        data[2], data[5], data[6], data[13], data[9], data[10], data[11], data[12]
    url = f'http://172.20.10.8:8000/client_card/{slug}/'
    create_qr(url)
    wb = load_workbook(f"{static_path}/media/excel_files/examples/warranty.xlsx")
    ws = wb.worksheets[0]
    short_date = date.strftime("%d.%m.%Y")
    longest_date = date.strftime("%d.%m.%Y %H:%M")
    string_width = 25
    ws["A2"], ws["A10"], ws["A12"], ws["A15"], ws["A8"] =\
        model, name, phone_num, short_date, warranty
    break_fix = textwrap.fill(break_fix, string_width).split("\n")
    ws["A4"] = break_fix[0]
    if len(break_fix) >= 2:
        ws["A5"] = break_fix[1]
    if len(break_fix) >= 3:                             
        ws["A6"] = break_fix[2]
    wb.save(f"{static_path}/media/excel_files/sawed_xlsx/{name}_{longest_date}warranty.xlsx")
    return f'media/excel_files/sawed_xlsx/{name}_{longest_date}warranty.xlsx'


if __name__ == '__main__':
    print(Path(__file__).resolve().parent)
